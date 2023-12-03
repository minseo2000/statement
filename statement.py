# -*- coding: utf-8 -*-
import openpyxl as op
import os
import datetime


# 전표 뼈대 생성 클래스   account = 거래처 이름,  phone_num 전화번호, location_place = 주소
class Statement:

    # 오늘 날짜
    today = datetime.datetime.now()
    save_path = today.strftime('%Y년 %m월 %d일'.encode('unicode-escape').decode()).encode().decode('unicode-escape')
    # 워크북 객체
    workbook = ""

    # sheet 객체
    intercell = ""

    def __init__(self, date_str, date, path, NAME, CHAIRMAN, CATEGORY, CATE, NUM, LOCATION):  # constructor
        print('생성완료!')
        self.NAME = NAME
        self.CHAIRMAN = CHAIRMAN
        self.CATEGORY = CATEGORY
        self.CATE = CATE
        self.NUM = NUM
        self.date_str = date_str
        self.date = date
        self.path = date + ".xlsx"
        self.LOCATION = LOCATION

    def getConst(self):
        print(self.account)
        print(self.phone_num)
        print(self.location_place)

    def make_new_workbook(self):
        
        if not os.path.exists(self.save_path):
            os.mkdir(self.save_path)
        else:
            print('이미 존재')
        
        isFile = False
        
        for i in os.listdir(os.getcwd() + "/"+ self.save_path+'/'):
            if i == self.path:
                isFile = True

        if isFile == True:
            self.workbook = op.load_workbook(filename=self.save_path+"/" + self.path)
        else:
            self.workbook = op.Workbook()

    # sheet 뼈대 만들기

    def make_new_layout(self):
        intercell = self.workbook.create_sheet()
        self.intercell = intercell
        # 거래명세서 뼈대 만들기   4, 6, 8, 10
        intercell.column_dimensions['A'].width = 3
        intercell.column_dimensions['B'].width = 3
        intercell.column_dimensions['C'].width = 9
        intercell.column_dimensions['D'].width = 5
        intercell.column_dimensions['E'].width = 3
        intercell.column_dimensions['F'].width = 7
        intercell.column_dimensions['G'].width = 3
        intercell.column_dimensions['H'].width = 7
        intercell.column_dimensions['I'].width = 5
        intercell.column_dimensions['J'].width = 7
        intercell.column_dimensions['K'].width = 9

        # 정렬 초기화
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in ["A", "B", "C", "D", 'E', "F", "G", "H", "I", "J", "K"]:
            for j in range(1, 37):
                intercell[i + str(j)].alignment = format1

                # 중간 분리대
        intercell.column_dimensions['L'].width = 5

        for i in range(1, 37):
            intercell.row_dimensions[i].height = 13

        intercell['G1'] = self.NUM
        intercell['G3'] = self.NAME
        intercell['C5'] = self.date_str
        intercell['K3'] = self.CHAIRMAN
        intercell['G5'] = self.LOCATION
        intercell['G7'] = self.CATEGORY
        intercell['K7'] = self.CATE

        intercell["A1"] = "거래명세서"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = op.styles.Font(size=22)
        intercell["A1"].alignment = format1
        intercell["A1"].font = font

        intercell["A4"] = "(공급받는자용)"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["A4"].alignment = format1

        intercell["A5"] = "거래일 :"
        font = op.styles.Font(size=9)
        intercell["A5"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["A5"].alignment = format1

        intercell["D7"] = "귀하"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["D7"].alignment = format1

        intercell["E1"] = "공\n\n급\n\n자"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["E1"].alignment = format1

        intercell["F1"] = "사업자\n번호"
        font = op.styles.Font(size=10)
        intercell["F1"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["F1"].alignment = format1

        intercell["F3"] = "상호\n(법인명)"
        font = op.styles.Font(size=10)
        intercell["F3"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["F3"].alignment = format1

        intercell["F5"] = "사업장\n주소"
        font = op.styles.Font(size=10)
        intercell["F5"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["F5"].alignment = format1

        intercell["F7"] = "업태"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["F7"].alignment = format1

        intercell["J3"] = "성명\n(대표자)"
        font = op.styles.Font(size=10)
        intercell["J3"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["J3"].alignment = format1

        intercell["J7"] = "종목"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["J7"].alignment = format1

        intercell["A10"] = "합계금액\n(부가세 포함)"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = op.styles.Font(size=10)
        intercell["A10"].font = font
        intercell["A10"].alignment = format1

        intercell["K10"] = "원"
        format1 = op.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        intercell["K10"].alignment = format1

        intercell["A12"] = "NO"
        format1 = op.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        font = op.styles.Font(size=9)
        intercell["A12"].font = font
        intercell["A12"].alignment = format1

        intercell["B12"] = "품 목"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["B12"].alignment = format1

        intercell["F12"] = "수 량"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["F12"].alignment = format1

        intercell["G12"] = "단 가"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["G12"].alignment = format1

        intercell["I12"] = "공급가액"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["I12"].alignment = format1

        intercell["K12"] = "세 액"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["K12"].alignment = format1

        intercell["A32"] = "계"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["A32"].alignment = format1

        intercell["A33"] = "비고"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["A33"].alignment = format1

        intercell["A35"] = "납품자"
        font = op.styles.Font(size=10)
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["A35"].alignment = format1
        intercell["A35"].font = font

        intercell["F35"] = "인수자"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["F35"].alignment = format1

        for i in range(1, 20):
            intercell["A" + str(i + 12)] = i
            format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            intercell["A" + str(i + 12)].alignment = format1

        intercell.merge_cells('A1:D3')  # 거래 명세서 부분

        intercell.merge_cells('A4:D4')  # 공급받는자 , 공급자 부분

        intercell.merge_cells('A5:B6')  # 거래일 문자열 부분

        intercell.merge_cells('C5:D6')  # 거래일 실제로 쓰는 부분

        intercell.merge_cells('A7:C8')  # 귀하 실제로 쓰는 부분

        intercell.merge_cells('D7:D8')  # 귀하 문자열 쓰는 부분

        intercell.merge_cells('A9:C9')  # 주소 이름 쓰는 부분

        intercell.merge_cells('D9:K9')  # 주소 실제로 쓰는 부분

        intercell.merge_cells('A10:C11')  # 합계금액(부가세 포함) 쓰는 부분

        intercell.merge_cells('D10:J11')  # 금액 쓰는 부분

        intercell.merge_cells('K10:K11')  # 원 글씨 쓰는 부분

        intercell.merge_cells('E1:E8')  # 공급자 글씨 쓰는 부분

        intercell.merge_cells('F1:F2')  # 사업자 번호 글씨 쓰는 부분

        intercell.merge_cells('F3:F4')  # 상호(법인명) 글씨 쓰는 부분

        intercell.merge_cells('F5:F6')  # 사업장 주소 글씨 쓰는 부분

        intercell.merge_cells('F7:F8')  # 업태 글씨 쓰는 부분

        intercell.merge_cells('G1:K2')  # 사업자 번호 실제로 쓰는 부분

        intercell.merge_cells('G3:I4')  # 상호 법인명 쓰는 부분

        intercell.merge_cells('J3:J4')  # 성명(대표자) 글씨 쓰는 부분

        intercell.merge_cells('K3:K4')  # 성명 실제로 쓰는 부분

        intercell.merge_cells('G5:K6')  # 사업장 주소 실제로 쓰는 부분

        intercell.merge_cells('G7:I8')  # 도 소매 글씨 쓰는 부분

        intercell.merge_cells('J7:J8')  # 종목 글씨 쓰는 부분

        intercell.merge_cells('K7:K8')  # 축전지 글씨 쓰는 부분

        intercell.merge_cells('B12:E12')  # 품목 부분

        for i in range(12, 32):
            estr = "B" + str(i) + ":" + "E" + str(i)
            intercell.merge_cells(estr)  # 품목 부분

        for i in range(12, 33):
            estr = "G" + str(i) + ":" + "H" + str(i)
            intercell.merge_cells(estr)  # 단가 부분
        for i in range(12, 33):
            estr = "I" + str(i) + ":" + "J" + str(i)
            intercell.merge_cells(estr)  # 공급가액 부분

        intercell.merge_cells('A32:E32')  # 계 글씨 쓰는 부분

        intercell.merge_cells('A33:B34')  # 비고 글씨 부분

        intercell.merge_cells('C33:K34')  # 비고 실제 쓰는 부분

        intercell.merge_cells('A35:B36')  # 납품자 글씨 쓰는부분

        intercell.merge_cells('C35:E36')  # 납품자 실제 쓰는부분

        intercell.merge_cells('F35:G36')  # 인수자 쓰는 부분

        intercell.merge_cells('H35:J36')  # 인수자 실제 쓰는부분

        intercell.merge_cells('K35:K36')  # 부분

        # 내부 보더 설정
        for i in ["A", "B", "C", "D", 'E', "F", "G", "H", "I", "J", "K"]:
            for j in range(1, 37):
                intercell[i + str(j)].border = op.styles.Border(top=op.styles.Side(border_style="dotted"),
                                                                bottom=op.styles.Side(border_style="dotted"),
                                                                left=op.styles.Side(border_style="dotted"),
                                                                right=op.styles.Side(border_style="dotted"))

        # 겉 보더 설정
        intercell["A1"].border = op.styles.Border(top=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="dotted"),
                                                  left=op.styles.Side(border_style="dotted"),
                                                  right=op.styles.Side(border_style="dotted"))

        for i in ["A", "B", "C", "D", 'E', "F", "G", "H", "I", "J", "K"]:
            intercell[i + "36"].border = op.styles.Border(bottom=op.styles.Side(border_style="thin"),
                                                          top=op.styles.Side(border_style="dotted"),
                                                          left=op.styles.Side(border_style="dotted"),
                                                          right=op.styles.Side(border_style="dotted"))
            intercell[i + "1"].border = op.styles.Border(top=op.styles.Side(border_style="thin"),
                                                         bottom=op.styles.Side(border_style="dotted"),
                                                         left=op.styles.Side(border_style="dotted"),
                                                         right=op.styles.Side(border_style="dotted"))

        for i in range(1, 37):
            intercell["K" + str(i)].border = op.styles.Border(right=op.styles.Side(border_style="thin"),
                                                              bottom=op.styles.Side(border_style="dotted"),
                                                              left=op.styles.Side(border_style="dotted"),
                                                              top=op.styles.Side(border_style="dotted"))
            intercell["A" + str(i)].border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                              bottom=op.styles.Side(border_style="dotted"),
                                                              top=op.styles.Side(border_style="dotted"),
                                                              right=op.styles.Side(border_style="dotted"))

        intercell["A1"].border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                  top=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="dotted"),
                                                  right=op.styles.Side(border_style="dotted"))
        intercell["K36"].border = op.styles.Border(bottom=op.styles.Side(border_style="thin"),
                                                   right=op.styles.Side(border_style="thin"),
                                                   top=op.styles.Side(border_style="dotted"),
                                                   left=op.styles.Side(border_style="dotted"), )
        intercell["K1"].border = op.styles.Border(top=op.styles.Side(border_style="thin"),
                                                  right=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="dotted"),
                                                  left=op.styles.Side(border_style="dotted"), )

        intercell["A36"].border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                   bottom=op.styles.Side(border_style="thin"),
                                                   top=op.styles.Side(border_style="dotted"),
                                                   right=op.styles.Side(border_style="dotted"))

        #################################################### 공급자용

        intercell.column_dimensions['M'].width = 3
        intercell.column_dimensions['N'].width = 3
        intercell.column_dimensions['O'].width = 9
        intercell.column_dimensions['P'].width = 5
        intercell.column_dimensions['Q'].width = 3
        intercell.column_dimensions['R'].width = 7
        intercell.column_dimensions['S'].width = 3
        intercell.column_dimensions['T'].width = 7
        intercell.column_dimensions['U'].width = 5
        intercell.column_dimensions['V'].width = 7
        intercell.column_dimensions['W'].width = 9

        # 정렬 초기화
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        for i in ["M", "N", "O", "P", 'Q', "R", "S", "T", "U", "V", "W"]:
            for j in range(1, 37):
                intercell[i + str(j)].alignment = format1

        for i in range(1, 37):
            intercell.row_dimensions[i].height = 13

        intercell["M1"] = "거래명세서"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = op.styles.Font(size=22)
        intercell["M1"].alignment = format1
        intercell["M1"].font = font

        intercell["M4"] = "(공급자용)"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["M4"].alignment = format1

        intercell["M5"] = "거래일 :"
        font = op.styles.Font(size=9)
        intercell["M5"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["M5"].alignment = format1

        intercell["P7"] = "귀하"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["P7"].alignment = format1

        intercell["Q1"] = "공\n\n급\n\n자"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["Q1"].alignment = format1

        intercell["R1"] = "사업자\n번호"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = op.styles.Font(size=10)
        intercell["R1"].font = font
        intercell["R1"].alignment = format1

        intercell["R3"] = "상호\n(법인명)"
        font = op.styles.Font(size=10)
        intercell["R3"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["R3"].alignment = format1

        intercell["R5"] = "사업장\n주소"
        font = op.styles.Font(size=10)
        intercell["R5"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["R5"].alignment = format1

        intercell["R7"] = "업태"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["R7"].alignment = format1

        intercell["V3"] = "성명\n(대표자)"
        font = op.styles.Font(size=10)
        intercell["V3"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["V3"].alignment = format1

        intercell["V7"] = "종목"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["V7"].alignment = format1

        intercell["M10"] = "합계금액\n(부가세 포함)"
        font = op.styles.Font(size=10)
        intercell["M10"].font = font
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["M10"].alignment = format1

        intercell["W10"] = "원"
        format1 = op.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        font = op.styles.Font(size=9)
        intercell["W10"].font = font
        intercell["W10"].alignment = format1

        intercell["M12"] = "NO"
        format1 = op.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        font = op.styles.Font(size=9)
        intercell["M12"].font = font
        intercell["M12"].alignment = format1

        intercell["N12"] = "품 목"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["N12"].alignment = format1

        intercell["R12"] = "수 량"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["R12"].alignment = format1

        intercell["S12"] = "단 가"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["S12"].alignment = format1

        intercell["U12"] = "공급가액"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["U12"].alignment = format1

        intercell["W12"] = "세 액"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["W12"].alignment = format1

        intercell["M32"] = "계"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["M32"].alignment = format1

        intercell["M33"] = "비고"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["M33"].alignment = format1

        intercell["M35"] = "납품자"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = op.styles.Font(size=10)
        intercell["M35"].font = font
        intercell["M35"].alignment = format1

        intercell["R35"] = "인수자"
        format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        intercell["R35"].alignment = format1

        intercell['S1'] = self.NUM
        intercell['S3'] = self.NAME
        intercell['O5'] = self.date_str
        intercell['W3'] = self.CHAIRMAN
        intercell['S5'] = self.LOCATION
        intercell['S7'] = self.CATEGORY
        intercell['W7'] = self.CATE

        for i in range(1, 20):
            intercell["M" + str(i + 12)] = i
            format1 = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            intercell["M" + str(i + 12)].alignment = format1

        intercell.merge_cells('M1:P3')  # 거래 명세서 부분

        intercell.merge_cells('M4:P4')  # 공급받는자 , 공급자 부분

        intercell.merge_cells('M5:N6')  # 거래일 문자열 부분

        intercell.merge_cells('O5:P6')  # 거래일 실제로 쓰는 부분

        intercell.merge_cells('M7:O8')  # 귀하 실제로 쓰는 부분

        intercell.merge_cells('P7:P8')  # 귀하 문자열 쓰는 부분

        intercell.merge_cells('M9:O9')  # 주소 이름 쓰는 부분

        intercell.merge_cells('P9:W9')  # 주소 실제로 쓰는 부분

        intercell.merge_cells('M10:O11')  # 합계금액(부가세 포함) 쓰는 부분

        intercell.merge_cells('P10:V11')  # 금액 쓰는 부분

        intercell.merge_cells('W10:W11')  # 원 글씨 쓰는 부분

        intercell.merge_cells('Q1:Q8')  # 공급자 글씨 쓰는 부분

        intercell.merge_cells('R1:R2')  # 사업자 번호 글씨 쓰는 부분

        intercell.merge_cells('R3:R4')  # 상호(법인명) 글씨 쓰는 부분

        intercell.merge_cells('R5:R6')  # 사업장 주소 글씨 쓰는 부분

        intercell.merge_cells('R7:R8')  # 업태 글씨 쓰는 부분

        intercell.merge_cells('S1:W2')  # 사업자 번호 실제로 쓰는 부분

        intercell.merge_cells('S3:U4')  # 상호 법인명 쓰는 부분

        intercell.merge_cells('V3:V4')  # 성명(대표자) 글씨 쓰는 부분

        intercell.merge_cells('W3:W4')  # 성명 실제로 쓰는 부분

        intercell.merge_cells('S5:W6')  # 사업장 주소 실제로 쓰는 부분

        intercell.merge_cells('S7:U8')  # 도 소매 글씨 쓰는 부분

        intercell.merge_cells('V7:V8')  # 종목 글씨 쓰는 부분

        intercell.merge_cells('W7:W8')  # 축전지 글씨 쓰는 부분

        intercell.merge_cells('N12:Q12')  # 품목 부분

        for i in range(12, 32):
            estr = "N" + str(i) + ":" + "Q" + str(i)
            intercell.merge_cells(estr)  # 품목 부분

        for i in range(12, 33):
            estr = "S" + str(i) + ":" + "T" + str(i)
            intercell.merge_cells(estr)  # 단가 부분
        for i in range(12, 33):
            estr = "U" + str(i) + ":" + "V" + str(i)
            intercell.merge_cells(estr)  # 공급가액 부분

        intercell.merge_cells('M32:Q32')  # 계 글씨 쓰는 부분

        intercell.merge_cells('M33:N34')  # 비고 글씨 부분

        intercell.merge_cells('O33:W34')  # 비고 실제 쓰는 부분

        intercell.merge_cells('M35:N36')  # 납품자 글씨 쓰는부분

        intercell.merge_cells('O35:Q36')  # 납품자 실제 쓰는부분

        intercell.merge_cells('R35:S36')  # 인수자 쓰는 부분

        intercell.merge_cells('T35:V36')  # 인수자 실제 쓰는부분

        intercell.merge_cells('W35:W36')  # 부분

        # 내부 보더 설정
        for i in ["M", "N", "O", "P", 'Q', "R", "S", "T", "U", "V", "W"]:
            for j in range(1, 37):
                intercell[i + str(j)].border = op.styles.Border(top=op.styles.Side(border_style="dotted"),
                                                                bottom=op.styles.Side(border_style="dotted"),
                                                                left=op.styles.Side(border_style="dotted"),
                                                                right=op.styles.Side(border_style="dotted"))

        # 겉 보더 설정
        intercell["M1"].border = op.styles.Border(top=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="dotted"),
                                                  left=op.styles.Side(border_style="dotted"),
                                                  right=op.styles.Side(border_style="dotted"))

        for i in ["M", "N", "O", "P", 'Q', "R", "S", "T", "U", "V", "W"]:
            intercell[i + "36"].border = op.styles.Border(bottom=op.styles.Side(border_style="thin"),
                                                          top=op.styles.Side(border_style="dotted"),
                                                          left=op.styles.Side(border_style="dotted"),
                                                          right=op.styles.Side(border_style="dotted"))
            intercell[i + "1"].border = op.styles.Border(top=op.styles.Side(border_style="thin"),
                                                         bottom=op.styles.Side(border_style="dotted"),
                                                         left=op.styles.Side(border_style="dotted"),
                                                         right=op.styles.Side(border_style="dotted"))

        for i in range(1, 37):
            intercell["W" + str(i)].border = op.styles.Border(right=op.styles.Side(border_style="thin"),
                                                              bottom=op.styles.Side(border_style="dotted"),
                                                              left=op.styles.Side(border_style="dotted"),
                                                              top=op.styles.Side(border_style="dotted"))
            intercell["M" + str(i)].border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                              bottom=op.styles.Side(border_style="dotted"),
                                                              top=op.styles.Side(border_style="dotted"),
                                                              right=op.styles.Side(border_style="dotted"))

        intercell["M1"].border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                  top=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="dotted"),
                                                  right=op.styles.Side(border_style="dotted"))
        intercell["W36"].border = op.styles.Border(bottom=op.styles.Side(border_style="thin"),
                                                   right=op.styles.Side(border_style="thin"),
                                                   top=op.styles.Side(border_style="dotted"),
                                                   left=op.styles.Side(border_style="dotted"), )
        intercell["W1"].border = op.styles.Border(top=op.styles.Side(border_style="thin"),
                                                  right=op.styles.Side(border_style="thin"),
                                                  bottom=op.styles.Side(border_style="dotted"),
                                                  left=op.styles.Side(border_style="dotted"), )

        intercell["M36"].border = op.styles.Border(left=op.styles.Side(border_style="thin"),
                                                   bottom=op.styles.Side(border_style="thin"),
                                                   top=op.styles.Side(border_style="dotted"),
                                                   right=op.styles.Side(border_style="dotted"))
        # intercell.set_printer_settings('8','landscape')
        intercell.print_area = "A1:W36"
        intercell.print_options.horizontalCentered = True
        intercell.print_options.verticalCentered = True
        intercell.page_setup.orientation = intercell.ORIENTATION_LANDSCAPE
        intercell.page_setup.paperSize = intercell.PAPERSIZE_A4
        self.workbook.save(os.getcwd() + "/"+self.save_path+'/'+ self.path)
        return intercell

    # account 매개변수는 거래처 정보! -> 딕셔너리 타입!
    # item 매개변수는 딕셔너리 타입!
    # 품목 칸 : B13~B31
    # 수량 칸 : F13~F31
    # 총 수량 : F32

    # 거래처 전화번호 칸 : A9
    # 거래처 주소 칸 : D9
    # 거래처 이름 칸 : A7

    # item dic
    # {name : DF90R}
    # {cnt : 1}

    # acoount dic
    # {name : 인터셀}
    # {ph_n : 01011111111}
    # {loc : 경기도 김포시}

    def enter_item(self, account, item, intercell):
        intercell.title = account['name']
        intercell['A7'] = account['name']
        intercell['M7'] = account['name']
        intercell['A9'] = account['ph_n']
        intercell['M9'] = account['ph_n']
        intercell['D9'] = account['loc']
        intercell['P9'] = account['loc']

        # 아이템 길이
        item_len = len(item['name'])

        for i in range(1, item_len + 1):
            intercell['B' + str(i + 12)] = item['name'][i - 1]  # 품목 저장
            intercell['F' + str(i + 12)] = item['cnt'][i - 1]
            intercell['G' + str(i + 12)] = item['price'][i - 1]
            intercell['N' + str(i + 12)] = item['name'][i - 1]  # 품목 저장
            intercell['R' + str(i + 12)] = item['cnt'][i - 1]
            intercell['S' + str(i + 12)] = item['price'][i - 1]

            print(item['price'])
        # 물건 수량 토탈 계산

        intercell['F32'] = "=SUM(F13:F31)"
        intercell['R32'] = "=SUM(F13:F31)"
        intercell['G32'] = "=SUM(G13:G31)"
        intercell['S32'] = "=SUM(G13:G31)"

        self.workbook.save(os.getcwd() + "/"+self.save_path+'/'+ self.path)

